import yfinance as yf
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
import os

curr_dir = os.getcwd()
EXCEL_FILE = os.path.join(curr_dir, 'State Bank of India.xlsx')

SHEET_NAME = 'Sheet1'
def fetch_friday_closing_price():
    ticker_symbol = 'SBIN.BO'
    sbi_data = yf.Ticker(ticker_symbol)
    hist = sbi_data.history(period='5d')
    fridays = hist[hist.index.weekday == 4]

    if not fridays.empty:
        friday_date = fridays.index[0].date()
        friday_close = fridays['Close'].iloc[0]
        
        if os.path.exists(EXCEL_FILE):
            book = load_workbook(EXCEL_FILE)
            if SHEET_NAME in book.sheetnames:
                df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
                df['DATE'] = pd.to_datetime(df['DATE']).dt.date
            else:
                df = pd.DataFrame(columns=['DATE', 'WEEKLY'])
        else:
            df = pd.DataFrame(columns=['DATE', 'WEEKLY'])

        if friday_date in df['DATE'].values:
            print(f"Data for {friday_date} is already present. Skipping appending.")
        else:
            new_row = pd.DataFrame({'DATE': [friday_date], 'WEEKLY': [friday_close]})
            df = pd.concat([df, new_row], ignore_index=True)
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
                
            book.save(EXCEL_FILE)
            book.close() 
            print(f"Appended data: {friday_date} - {friday_close}")
            send_email(EXCEL_FILE)
    else:
        print(f"No trading data available for today, {friday_date}")

def send_email(file_path):
    sender_email = os.getenv('SENDER_EMAIL')
    receiver_email = os.getenv('RECEIVER_EMAIL')
    password = os.getenv('EMAIL_PASSWORD')
    subject = 'SBI Friday Closing Prices'
    body = 'Please find attached the latest SBI Friday closing prices.'
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    attachment = MIMEBase('application', 'octet-stream')
    with open(file_path, 'rb') as attachment_file:
        attachment.set_payload(attachment_file.read())
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
    msg.attach(attachment)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)  # Use Gmail's SMTP server
        server.starttls()
        server.login(sender_email, password)
        text = msg.as_string()
        server.sendmail(sender_email, receiver_email, text)
        server.quit()
        print('Email sent successfully!')
    except Exception as e:
        print(f'Failed to send email: {e}')

fetch_friday_closing_price()
print("Task is done.")

